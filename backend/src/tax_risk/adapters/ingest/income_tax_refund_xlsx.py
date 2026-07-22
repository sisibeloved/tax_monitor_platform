"""Strict parser for the controlled annual income-tax refund workbook."""

from __future__ import annotations

from dataclasses import dataclass
from collections.abc import Iterable, Sequence
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Protocol, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

from tax_risk.adapters.ingest.base import fits_database_amount
from tax_risk.adapters.ingest.tax_master_xlsx import (
    DEFAULT_XLSX_RESOURCE_LIMITS,
    TaxMasterWorkbookError,
    XlsxResourceLimits,
    preflight_xlsx,
)
from tax_risk.domain.money import Money


_REFUND_WORKBOOK_AMOUNT_QUANTUM = Decimal("0.0001")


@dataclass(frozen=True, slots=True)
class IncomeTaxRefundWorkbookRow:
    row_number: int
    source_record_key: str
    unified_credit_code: str | None
    company_code: str
    company_name: str
    refund_tax_year: int
    raw_expected_refund_amount: Decimal
    expected_refund_amount: Decimal
    received_in_source: bool


@dataclass(frozen=True, slots=True)
class IncomeTaxRefundWorkbookRowError:
    row_number: int
    error_code: str
    message: str
    field: str | None = None


class IncomeTaxRefundWorkbookError(ValueError):
    def __init__(self, errors: tuple[IncomeTaxRefundWorkbookRowError, ...]) -> None:
        if not errors:
            raise ValueError("workbook error requires at least one row error")
        self.errors = errors
        super().__init__(errors[0].message)


class _Worksheet(Protocol):
    def iter_rows(
        self,
        *,
        min_row: int | None = None,
        max_row: int | None = None,
        values_only: bool = False,
    ) -> Iterable[tuple[object, ...]]: ...


class IncomeTaxRefundXlsxAdapter:
    """Read refund targets from the latest controlled legal-entity workbook."""

    def __init__(
        self,
        payload: bytes,
        *,
        refund_tax_year: int,
        currency: str = "CNY",
        amount_scale: int = 2,
        max_upload_bytes: int = 16 * 1024 * 1024,
        limits: XlsxResourceLimits = DEFAULT_XLSX_RESOURCE_LIMITS,
    ) -> None:
        if not isinstance(payload, bytes):
            raise TypeError("payload must be bytes")
        if not 2000 <= refund_tax_year <= 9998:
            raise ValueError("refund_tax_year must be between 2000 and 9998")
        if type(amount_scale) is not int or not 0 <= amount_scale <= 12:
            raise ValueError("amount_scale must be between 0 and 12")
        if type(max_upload_bytes) is not int or max_upload_bytes <= 0:
            raise ValueError("max_upload_bytes must be positive")
        if not isinstance(limits, XlsxResourceLimits):
            raise TypeError("limits must be XlsxResourceLimits")
        normalized_currency = currency.strip().upper()
        if len(normalized_currency) != 3 or not normalized_currency.isalpha():
            raise ValueError("currency must be a three-letter code")
        if len(payload) > max_upload_bytes:
            raise IncomeTaxRefundWorkbookError(
                (
                    IncomeTaxRefundWorkbookRowError(
                        1,
                        "XLSX_RESOURCE_LIMIT_EXCEEDED",
                        "refund workbook exceeds the upload-size limit",
                    ),
                )
            )
        self._payload = payload
        self._refund_tax_year = refund_tax_year
        self._currency = normalized_currency
        self._amount_scale = amount_scale
        self._limits = limits

    def parse(self) -> tuple[IncomeTaxRefundWorkbookRow, ...]:
        try:
            preflight_xlsx(self._payload, self._limits)
            workbook = load_workbook(
                BytesIO(self._payload),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except TaxMasterWorkbookError as error:
            raise IncomeTaxRefundWorkbookError(
                tuple(
                    IncomeTaxRefundWorkbookRowError(
                        item.row_number,
                        item.error_code,
                        item.message,
                        item.field,
                    )
                    for item in error.errors
                )
            ) from error
        except Exception as error:
            raise IncomeTaxRefundWorkbookError(
                (
                    IncomeTaxRefundWorkbookRowError(
                        1,
                        "INVALID_XLSX",
                        "file is not a readable XLSX workbook",
                    ),
                )
            ) from error
        try:
            worksheet, header = self._locate_sheet(
                cast(Sequence[_Worksheet], workbook.worksheets)
            )
            return self._parse_rows(worksheet.iter_rows(values_only=True), header)
        finally:
            workbook.close()

    def _locate_sheet(
        self,
        worksheets: Sequence[_Worksheet],
    ) -> tuple[_Worksheet, dict[str, int]]:
        required = self._required_headers()
        for worksheet in worksheets:
            first_row = next(
                iter(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)),
                (),
            )
            header = {
                normalized: index
                for index, value in enumerate(first_row)
                if (normalized := _text(value)) is not None
            }
            if required <= header.keys():
                return worksheet, header
        missing = ", ".join(sorted(required))
        raise IncomeTaxRefundWorkbookError(
            (
                IncomeTaxRefundWorkbookRowError(
                    1,
                    "INVALID_HEADER",
                    f"no worksheet contains the required refund columns: {missing}",
                ),
            )
        )

    def _parse_rows(
        self,
        rows: Iterable[tuple[object, ...]],
        header: dict[str, int],
    ) -> tuple[IncomeTaxRefundWorkbookRow, ...]:
        iterator = iter(rows)
        next(iterator, None)
        parsed: list[IncomeTaxRefundWorkbookRow] = []
        errors: list[IncomeTaxRefundWorkbookRowError] = []
        seen_companies: set[str] = set()
        for row_number, values in enumerate(iterator, start=2):
            row = tuple(values)
            involved = _text(_at(row, header[self._involved_header]))
            if involved is None and all(value in (None, "") for value in row):
                continue
            if involved not in {"是", "否"}:
                errors.append(
                    IncomeTaxRefundWorkbookRowError(
                        row_number,
                        "INVALID_REFUND_FLAG",
                        f"{self._involved_header} must be 是 or 否",
                        self._involved_header,
                    )
                )
                continue
            if involved == "否":
                continue

            company_code = _identifier(
                _at(row, header["公司代码"]),
                row_number=row_number,
                field="公司代码",
                errors=errors,
            )
            company_name = _required_text(
                _at(row, header["公司名称"]),
                row_number=row_number,
                field="公司名称",
                errors=errors,
            )
            unified_credit_code = (
                _identifier(
                    _at(row, header["统一信用代码"]),
                    row_number=row_number,
                    field="统一信用代码",
                    errors=errors,
                    optional=True,
                )
                if "统一信用代码" in header
                else None
            )
            raw_amount = _positive_decimal(
                _at(row, header[self._amount_header]),
                row_number=row_number,
                field=self._amount_header,
                errors=errors,
            )
            source_status = _text(_at(row, header["是否已收到退税"]))
            if source_status not in {None, "未退税", "已退税"}:
                errors.append(
                    IncomeTaxRefundWorkbookRowError(
                        row_number,
                        "INVALID_RECEIPT_STATUS",
                        "是否已收到退税 must be blank, 未退税, or 已退税",
                        "是否已收到退税",
                    )
                )
            if company_code is None or company_name is None or raw_amount is None:
                continue
            if company_code in seen_companies:
                errors.append(
                    IncomeTaxRefundWorkbookRowError(
                        row_number,
                        "DUPLICATE_COMPANY_REFUND",
                        "one company may have only one refund target per tax year",
                        "公司代码",
                    )
                )
                continue
            seen_companies.add(company_code)
            normalized_amount = Money.unrounded(
                raw_amount,
                currency=self._currency,
                scale=self._amount_scale,
            ).quantized().amount
            if not fits_database_amount(normalized_amount):
                errors.append(
                    IncomeTaxRefundWorkbookRowError(
                        row_number,
                        "AMOUNT_OVERFLOW",
                        "expected refund amount does not fit NUMERIC(38, 12)",
                        self._amount_header,
                    )
                )
                continue
            parsed.append(
                IncomeTaxRefundWorkbookRow(
                    row_number=row_number,
                    source_record_key=(
                        f"{unified_credit_code or company_code}|"
                        f"{company_code}|{self._refund_tax_year}"
                    ),
                    unified_credit_code=unified_credit_code,
                    company_code=company_code,
                    company_name=company_name,
                    refund_tax_year=self._refund_tax_year,
                    raw_expected_refund_amount=raw_amount,
                    expected_refund_amount=normalized_amount,
                    received_in_source=source_status == "已退税",
                )
            )
        if errors:
            errors.sort(key=lambda item: (item.row_number, item.field or "", item.error_code))
            raise IncomeTaxRefundWorkbookError(tuple(errors))
        return tuple(parsed)

    @property
    def _involved_header(self) -> str:
        return f"{self._refund_tax_year}年是否涉及退税"

    @property
    def _amount_header(self) -> str:
        return f"{self._refund_tax_year}年应退税金额"

    def _required_headers(self) -> set[str]:
        return {
            "公司代码",
            "公司名称",
            self._involved_header,
            self._amount_header,
            "是否已收到退税",
        }


def _at(row: tuple[object, ...], index: int) -> object:
    return row[index] if index < len(row) else None


def _text(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    return normalized or None


def _required_text(
    value: object,
    *,
    row_number: int,
    field: str,
    errors: list[IncomeTaxRefundWorkbookRowError],
) -> str | None:
    normalized = _text(value)
    if normalized is None:
        errors.append(
            IncomeTaxRefundWorkbookRowError(
                row_number,
                "MISSING_VALUE",
                f"{field} is required",
                field,
            )
        )
    return normalized


def _identifier(
    value: object,
    *,
    row_number: int,
    field: str,
    errors: list[IncomeTaxRefundWorkbookRowError],
    optional: bool = False,
) -> str | None:
    normalized: str | None
    if isinstance(value, str):
        normalized = value.strip() or None
    elif type(value) is int:
        normalized = str(value)
    elif type(value) is float and value.is_integer():
        normalized = str(int(value))
    else:
        normalized = None
    if normalized is None and not optional:
        errors.append(
            IncomeTaxRefundWorkbookRowError(
                row_number,
                "MISSING_VALUE",
                f"{field} is required",
                field,
            )
        )
    elif normalized is not None and len(normalized) > 64:
        errors.append(
            IncomeTaxRefundWorkbookRowError(
                row_number,
                "VALUE_TOO_LONG",
                f"{field} must not exceed 64 characters",
                field,
            )
        )
        return None
    return normalized


def _positive_decimal(
    value: object,
    *,
    row_number: int,
    field: str,
    errors: list[IncomeTaxRefundWorkbookRowError],
) -> Decimal | None:
    try:
        if isinstance(value, bool) or value is None:
            raise ValueError
        if isinstance(value, Decimal):
            amount = value
        elif isinstance(value, str):
            amount = Decimal(value.strip().replace(",", ""))
        elif type(value) is int:
            amount = Decimal(str(value))
        elif type(value) is float:
            amount = Decimal(str(value)).quantize(
                _REFUND_WORKBOOK_AMOUNT_QUANTUM,
                rounding=ROUND_HALF_UP,
            )
        else:
            raise ValueError
        if not amount.is_finite() or amount <= 0 or not fits_database_amount(amount):
            raise ValueError
        return amount
    except (InvalidOperation, TypeError, ValueError):
        errors.append(
            IncomeTaxRefundWorkbookRowError(
                row_number,
                "INVALID_REFUND_AMOUNT",
                f"{field} must be a positive finite amount",
                field,
            )
        )
        return None


__all__ = [
    "IncomeTaxRefundWorkbookError",
    "IncomeTaxRefundWorkbookRow",
    "IncomeTaxRefundWorkbookRowError",
    "IncomeTaxRefundXlsxAdapter",
]
