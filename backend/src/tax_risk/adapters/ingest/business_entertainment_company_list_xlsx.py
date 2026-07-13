"""Strict XLSX adapter for the controlled business-entertainment company scope."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO

from openpyxl import load_workbook  # type: ignore[import-untyped]


REQUIRED_COLUMNS = ("company_code", "effective_from", "effective_to")


@dataclass(frozen=True, slots=True)
class BusinessEntertainmentScopeRow:
    row_number: int
    company_code: str
    effective_from: date
    effective_to: date


@dataclass(frozen=True, slots=True)
class BusinessEntertainmentScopeRowError:
    row_number: int
    error_code: str
    message: str
    field: str | None = None
    rejected_value: str | None = None


class BusinessEntertainmentScopeWorkbookError(ValueError):
    def __init__(
        self,
        errors: tuple[BusinessEntertainmentScopeRowError, ...],
        *,
        record_count: int = 0,
    ) -> None:
        if not errors:
            raise ValueError("workbook error requires at least one row error")
        self.errors = errors
        self.record_count = record_count
        super().__init__(errors[0].message)


class BusinessEntertainmentScopeXlsxAdapter:
    def __init__(self, payload: bytes) -> None:
        if not isinstance(payload, bytes):
            raise TypeError("payload must be bytes")
        self._payload = payload

    @property
    def checksum(self) -> str:
        return sha256(self._payload).hexdigest()

    def parse(self) -> tuple[BusinessEntertainmentScopeRow, ...]:
        try:
            workbook = load_workbook(
                BytesIO(self._payload),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as error:
            raise BusinessEntertainmentScopeWorkbookError(
                (
                    BusinessEntertainmentScopeRowError(
                        row_number=1,
                        error_code="INVALID_XLSX",
                        message="file is not a readable XLSX workbook",
                    ),
                )
            ) from error

        try:
            worksheet = workbook.active
            worksheet.reset_dimensions()
            return _parse_rows(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()


def _parse_rows(
    raw_rows: Iterable[tuple[object, ...]],
) -> tuple[BusinessEntertainmentScopeRow, ...]:
    rows = iter(raw_rows)
    header = next(rows, None)
    if header is None or tuple(header) != REQUIRED_COLUMNS:
        raise BusinessEntertainmentScopeWorkbookError(
            (
                BusinessEntertainmentScopeRowError(
                    row_number=1,
                    error_code="INVALID_HEADER",
                    message="header must contain exactly the approved scope columns",
                ),
            )
        )

    parsed: list[BusinessEntertainmentScopeRow] = []
    errors: list[BusinessEntertainmentScopeRowError] = []
    seen_companies: dict[str, int] = {}
    record_count = 0
    for row_number, values in enumerate(rows, start=2):
        if _is_blank(values):
            continue
        record_count += 1
        company_code = _company_code(values[0] if values else None, row_number, errors)
        effective_from = _date_value(
            values[1] if len(values) > 1 else None,
            row_number,
            "effective_from",
            errors,
        )
        effective_to = _date_value(
            values[2] if len(values) > 2 else None,
            row_number,
            "effective_to",
            errors,
        )
        if company_code is not None:
            previous_row = seen_companies.get(company_code)
            if previous_row is not None:
                errors.append(
                    BusinessEntertainmentScopeRowError(
                        row_number=row_number,
                        error_code="DUPLICATE_COMPANY",
                        message=f"company_code duplicates row {previous_row}",
                        field="company_code",
                        rejected_value=company_code,
                    )
                )
            else:
                seen_companies[company_code] = row_number
        if (
            effective_from is not None
            and effective_to is not None
            and effective_to < effective_from
        ):
            errors.append(
                BusinessEntertainmentScopeRowError(
                    row_number=row_number,
                    error_code="INVALID_EFFECTIVE_PERIOD",
                    message="effective_to must be on or after effective_from",
                    field="effective_to",
                    rejected_value=effective_to.isoformat(),
                )
            )
        if company_code is not None and effective_from is not None and effective_to is not None:
            parsed.append(
                BusinessEntertainmentScopeRow(
                    row_number=row_number,
                    company_code=company_code,
                    effective_from=effective_from,
                    effective_to=effective_to,
                )
            )

    if not parsed and not errors:
        errors.append(
            BusinessEntertainmentScopeRowError(
                row_number=1,
                error_code="EMPTY_FILE",
                message="workbook contains no scope rows",
            )
        )
    periods = {(row.effective_from, row.effective_to) for row in parsed}
    if len(periods) > 1:
        errors.append(
            BusinessEntertainmentScopeRowError(
                row_number=1,
                error_code="INCONSISTENT_EFFECTIVE_PERIOD",
                message="all scope rows must use one effective period",
                field="effective_from",
            )
        )
    if errors:
        errors.sort(key=lambda issue: (issue.row_number, issue.field or "", issue.error_code))
        raise BusinessEntertainmentScopeWorkbookError(
            tuple(errors),
            record_count=record_count,
        )
    return tuple(parsed)


def _company_code(
    value: object,
    row_number: int,
    errors: list[BusinessEntertainmentScopeRowError],
) -> str | None:
    if not isinstance(value, str) or not value.strip():
        errors.append(
            BusinessEntertainmentScopeRowError(
                row_number=row_number,
                error_code="REQUIRED_VALUE",
                message="company_code is required",
                field="company_code",
                rejected_value=None if value is None else str(value),
            )
        )
        return None
    normalized = value.strip()
    if len(normalized) > 64:
        errors.append(
            BusinessEntertainmentScopeRowError(
                row_number=row_number,
                error_code="VALUE_TOO_LONG",
                message="company_code must be at most 64 characters",
                field="company_code",
                rejected_value=normalized,
            )
        )
        return None
    return normalized


def _date_value(
    value: object,
    row_number: int,
    field: str,
    errors: list[BusinessEntertainmentScopeRowError],
) -> date | None:
    parsed: date | None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = date.fromisoformat(value.strip())
        except ValueError:
            parsed = None
    else:
        parsed = None
    if parsed is None:
        errors.append(
            BusinessEntertainmentScopeRowError(
                row_number=row_number,
                error_code="REQUIRED_VALUE" if value in (None, "") else "INVALID_DATE",
                message=f"{field} must be an ISO date",
                field=field,
                rejected_value=None if value is None else str(value),
            )
        )
    return parsed


def _is_blank(values: tuple[object, ...]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


__all__ = [
    "BusinessEntertainmentScopeRow",
    "BusinessEntertainmentScopeRowError",
    "BusinessEntertainmentScopeWorkbookError",
    "BusinessEntertainmentScopeXlsxAdapter",
]
