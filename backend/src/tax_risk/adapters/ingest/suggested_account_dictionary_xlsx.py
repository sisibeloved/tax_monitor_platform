"""Strict XLSX adapter for the shared suggested-account dictionary."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from tax_risk.domain.semantic.account_dictionary import AccountEntryStatus


REQUIRED_COLUMNS = (
    "dictionary_version",
    "account_id",
    "account_code",
    "account_name",
    "accounting_classification",
    "allowed_monitor_types",
    "allowed_labels",
    "effective_from",
    "effective_to",
    "status",
)
REQUIRED_PHASE_2_ACCOUNTS = frozenset(
    {"MEETING_EXPENSE", "EMPLOYEE_EDUCATION", "EMPLOYEE_WELFARE", "MANUAL_REVIEW"}
)


@dataclass(frozen=True, slots=True)
class SuggestedAccountRow:
    row_number: int
    dictionary_version: str
    account_id: str
    account_code: str
    account_name: str
    accounting_classification: str
    allowed_monitor_types: tuple[str, ...]
    allowed_labels: tuple[str, ...]
    effective_from: date
    effective_to: date
    status: AccountEntryStatus


class SuggestedAccountWorkbookError(ValueError):
    pass


class SuggestedAccountDictionaryXlsxAdapter:
    def __init__(self, payload: bytes) -> None:
        if not isinstance(payload, bytes):
            raise TypeError("payload must be bytes")
        self._payload = payload

    @property
    def checksum(self) -> str:
        return sha256(self._payload).hexdigest()

    def parse(self) -> tuple[SuggestedAccountRow, ...]:
        try:
            workbook = load_workbook(
                BytesIO(self._payload),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as error:
            raise SuggestedAccountWorkbookError("file is not a readable XLSX workbook") from error
        try:
            worksheet = workbook.active
            worksheet.reset_dimensions()
            return _parse_rows(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()


def _parse_rows(raw_rows: Iterable[tuple[object, ...]]) -> tuple[SuggestedAccountRow, ...]:
    rows = iter(raw_rows)
    if tuple(next(rows, ())) != REQUIRED_COLUMNS:
        raise SuggestedAccountWorkbookError(
            "header must contain exactly the approved account dictionary columns"
        )
    parsed: list[SuggestedAccountRow] = []
    seen_ids: set[str] = set()
    for row_number, values in enumerate(rows, start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        if len(values) != len(REQUIRED_COLUMNS):
            raise SuggestedAccountWorkbookError(f"row {row_number} has an invalid shape")
        strings = tuple(_required_text(value, row_number) for value in values[:7])
        account_id = strings[1]
        if account_id in seen_ids:
            raise SuggestedAccountWorkbookError(f"row {row_number} duplicates account_id")
        seen_ids.add(account_id)
        effective_from = _date_value(values[7], row_number)
        effective_to = _date_value(values[8], row_number)
        if effective_to < effective_from:
            raise SuggestedAccountWorkbookError(f"row {row_number} has invalid effective period")
        try:
            status = AccountEntryStatus(_required_text(values[9], row_number))
        except ValueError as error:
            raise SuggestedAccountWorkbookError(f"row {row_number} has invalid status") from error
        parsed.append(
            SuggestedAccountRow(
                row_number=row_number,
                dictionary_version=strings[0],
                account_id=account_id,
                account_code=strings[2],
                account_name=strings[3],
                accounting_classification=strings[4],
                allowed_monitor_types=_csv_set(strings[5], row_number),
                allowed_labels=_csv_set(strings[6], row_number),
                effective_from=effective_from,
                effective_to=effective_to,
                status=status,
            )
        )
    if not parsed:
        raise SuggestedAccountWorkbookError("workbook contains no account rows")
    versions = {row.dictionary_version for row in parsed}
    periods = {(row.effective_from, row.effective_to) for row in parsed}
    if len(versions) != 1 or len(periods) != 1:
        raise SuggestedAccountWorkbookError("all rows must share one version and effective period")
    missing = REQUIRED_PHASE_2_ACCOUNTS - {row.account_id for row in parsed}
    if missing:
        raise SuggestedAccountWorkbookError(
            f"required phase-2 accounts are missing: {','.join(sorted(missing))}"
        )
    return tuple(sorted(parsed, key=lambda row: row.account_id))


def _required_text(value: object, row_number: int) -> str:
    if value is None or not str(value).strip():
        raise SuggestedAccountWorkbookError(f"row {row_number} contains a blank required value")
    return str(value).strip()


def _date_value(value: object, row_number: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as error:
        raise SuggestedAccountWorkbookError(f"row {row_number} contains an invalid date") from error


def _csv_set(value: str, row_number: int) -> tuple[str, ...]:
    members = tuple(member.strip() for member in value.split(","))
    if any(not member for member in members) or len(members) != len(set(members)):
        raise SuggestedAccountWorkbookError(f"row {row_number} contains an invalid list")
    return members


__all__ = [
    "REQUIRED_COLUMNS",
    "SuggestedAccountDictionaryXlsxAdapter",
    "SuggestedAccountRow",
    "SuggestedAccountWorkbookError",
]
