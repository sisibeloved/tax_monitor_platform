"""Strict, all-or-nothing adapter for controlled tax-master workbooks."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
import re
from typing import IO, Protocol, cast
from urllib.parse import urlsplit
from xml.etree.ElementTree import Element, ParseError, fromstring, iterparse
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]

from tax_risk.adapters.ingest.base import fits_database_amount
from tax_risk.domain.money import Money, Rate


REQUIRED_COLUMNS = (
    "company_code",
    "company_name",
    "valid_from",
    "valid_to",
    "tax_rate",
    "deferred_tax_rate",
    "loss_carryforward",
    "three_year_average_tax_burden",
)


@dataclass(frozen=True, slots=True)
class XlsxResourceLimits:
    max_zip_members: int = 128
    max_total_uncompressed_bytes: int = 64 * 1024 * 1024
    max_member_uncompressed_bytes: int = 32 * 1024 * 1024
    max_compression_ratio: int = 200
    max_worksheet_rows: int = 20_000
    max_worksheet_cells: int = 200_000

    def __post_init__(self) -> None:
        for field_name in self.__dataclass_fields__:
            value = getattr(self, field_name)
            if type(value) is not int or value <= 0:
                raise ValueError(f"{field_name} must be a positive integer")


DEFAULT_XLSX_RESOURCE_LIMITS = XlsxResourceLimits()
_CONTENT_TYPES_MEMBER = "[Content_Types].xml"
_WORKBOOK_MEMBER = "xl/workbook.xml"
_WORKBOOK_RELATIONSHIPS_MEMBER = "xl/_rels/workbook.xml.rels"
_WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
_WORKSHEET_RELATIONSHIP_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
)
_CHARTSHEET_RELATIONSHIP_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chartsheet"
)
_BOUNDED_CELL_REFERENCE = re.compile(r"([A-Z]{1,3})([1-9][0-9]{0,6})")
_BOUNDED_ROW_REFERENCE = re.compile(r"[1-9][0-9]{0,6}")
_MAX_CELL_REFERENCE_LENGTH = 10
_MAX_ROW_REFERENCE_LENGTH = 7
_EXCEL_MAX_COLUMN = 16_384
_EXCEL_MAX_ROW = 1_048_576


class _Cell(Protocol):
    value: object
    data_type: str


@dataclass(frozen=True, slots=True)
class TaxMasterRow:
    row_number: int
    company_code: str
    company_name: str
    valid_from: date
    valid_to: date | None
    tax_rate: Rate
    deferred_tax_rate: Rate
    loss_carryforward: Decimal
    three_year_average_tax_burden: Rate


@dataclass(frozen=True, slots=True)
class TaxMasterRowError:
    row_number: int
    error_code: str
    message: str
    field: str | None = None
    rejected_value: str | None = None


class TaxMasterWorkbookError(ValueError):
    def __init__(
        self,
        errors: tuple[TaxMasterRowError, ...],
        *,
        record_count: int = 0,
        valid_rows: tuple[TaxMasterRow, ...] = (),
        loss_control_total: Decimal | None = None,
    ) -> None:
        if not errors:
            raise ValueError("workbook error requires at least one row error")
        self.errors = errors
        self.record_count = record_count
        self.valid_rows = valid_rows
        self.loss_control_total = loss_control_total
        super().__init__(errors[0].message)


class TaxMasterXlsxAdapter:
    """Read the active worksheet with values and a parallel formula audit stream."""

    def __init__(
        self,
        payload: bytes,
        *,
        amount_scale: int,
        limits: XlsxResourceLimits = DEFAULT_XLSX_RESOURCE_LIMITS,
    ) -> None:
        if not isinstance(payload, bytes):
            raise TypeError("payload must be bytes")
        if type(amount_scale) is not int or not 0 <= amount_scale <= 12:
            raise ValueError("amount_scale must be an integer between 0 and 12")
        if not isinstance(limits, XlsxResourceLimits):
            raise TypeError("limits must be XlsxResourceLimits")
        self._payload = payload
        self._amount_scale = amount_scale
        self._limits = limits

    @property
    def checksum(self) -> str:
        return sha256(self._payload).hexdigest()

    def parse(self) -> tuple[TaxMasterRow, ...]:
        try:
            _preflight_xlsx(self._payload, self._limits)
            formula_cells = self._read_formula_cells()
            return self._read_values(formula_cells)
        except TaxMasterWorkbookError:
            raise
        except Exception as error:
            raise TaxMasterWorkbookError(
                (
                    TaxMasterRowError(
                        1,
                        "INVALID_XLSX",
                        "file is not a readable XLSX workbook",
                    ),
                )
            ) from error

    def _read_formula_cells(self) -> dict[int, dict[int, object]]:
        workbook = load_workbook(
            BytesIO(self._payload),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        try:
            formulas: dict[int, dict[int, object]] = {}
            worksheet = workbook.active
            worksheet.reset_dimensions()
            for row_number, row in enumerate(worksheet.iter_rows(), start=1):
                for column_number, cell in enumerate(row, start=1):
                    value = cell.value
                    if cell.data_type == "f" or (
                        isinstance(value, str) and value.startswith("=")
                    ):
                        formulas.setdefault(row_number, {})[column_number] = value
            return formulas
        finally:
            workbook.close()

    def _read_values(
        self,
        formula_cells: dict[int, dict[int, object]],
    ) -> tuple[TaxMasterRow, ...]:
        workbook = load_workbook(
            BytesIO(self._payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            worksheet = workbook.active
            worksheet.reset_dimensions()
            return self._parse_sheet(worksheet.iter_rows(), formula_cells)
        finally:
            workbook.close()

    def _parse_sheet(
        self,
        value_rows: Iterable[tuple[_Cell, ...]],
        formula_cells: dict[int, dict[int, object]],
    ) -> tuple[TaxMasterRow, ...]:
        rows = iter(value_rows)
        value_header = next(rows, None)
        if value_header is None:
            raise _workbook_error(1, "INVALID_HEADER", "required header row is missing")
        header = tuple(cell.value for cell in value_header)
        if header != REQUIRED_COLUMNS:
            raise _workbook_error(
                1,
                "INVALID_HEADER",
                "header must contain exactly the required columns in the approved order",
            )

        parsed: list[TaxMasterRow] = []
        errors: list[TaxMasterRowError] = []
        pending_blank_rows: list[int] = []
        loss_values: list[Decimal] = []
        loss_control_available = True
        record_count = 0
        for row_number, value_cells in enumerate(rows, start=2):
            values = tuple(cell.value for cell in value_cells)
            row_formulas = formula_cells.get(row_number, {})
            if _all_blank(values) and not row_formulas:
                pending_blank_rows.append(row_number)
                continue
            record_count += 1
            loss_errors: list[TaxMasterRowError] = []
            control_loss = _parse_loss(
                values[6] if len(values) > 6 else None,
                row_number,
                self._amount_scale,
                loss_errors,
            )
            if control_loss is None:
                loss_control_available = False
            else:
                loss_values.append(control_loss)
            errors.extend(
                TaxMasterRowError(
                    blank_row,
                    "BLANK_ROW",
                    "blank rows are allowed only after the final data row",
                )
                for blank_row in pending_blank_rows
            )
            pending_blank_rows.clear()

            formula_errors = _formula_errors(row_number, value_cells, row_formulas)
            if formula_errors:
                errors.extend(formula_errors)
                continue
            row, row_errors = self._parse_row(row_number, value_cells)
            errors.extend(row_errors)
            if row is not None:
                parsed.append(row)

        if not parsed and not errors:
            raise _workbook_error(1, "EMPTY_FILE", "workbook contains no data rows")
        errors.extend(_period_overlap_errors(parsed))
        if errors:
            errors.sort(key=lambda error: (error.row_number, error.field or "", error.error_code))
            raise TaxMasterWorkbookError(
                tuple(errors),
                record_count=record_count,
                valid_rows=tuple(parsed),
                loss_control_total=(
                    _sum_losses(loss_values, self._amount_scale)
                    if loss_control_available
                    else None
                ),
            )
        return tuple(parsed)

    def _parse_row(
        self,
        row_number: int,
        cells: tuple[_Cell, ...],
    ) -> tuple[TaxMasterRow | None, list[TaxMasterRowError]]:
        values = tuple(cell.value for cell in cells)
        if len(values) != len(REQUIRED_COLUMNS):
            return None, [
                TaxMasterRowError(
                    row_number,
                    "INVALID_ROW_WIDTH",
                    "data row must contain exactly the required columns",
                )
            ]
        errors: list[TaxMasterRowError] = []
        company_code = _parse_text(values[0], row_number, "company_code", 64, errors)
        company_name = _parse_text(values[1], row_number, "company_name", 256, errors)
        valid_from = _parse_date(values[2], row_number, "valid_from", False, errors)
        valid_to = _parse_date(values[3], row_number, "valid_to", True, errors)
        tax_rate = _parse_rate(values[4], row_number, "tax_rate", errors)
        deferred_tax_rate = _parse_rate(
            values[5],
            row_number,
            "deferred_tax_rate",
            errors,
        )
        loss = _parse_loss(
            values[6],
            row_number,
            self._amount_scale,
            errors,
        )
        burden = _parse_rate(
            values[7],
            row_number,
            "three_year_average_tax_burden",
            errors,
        )
        if valid_from is not None and valid_to is not None and valid_to < valid_from:
            errors.append(
                TaxMasterRowError(
                    row_number,
                    "INVALID_EFFECTIVE_PERIOD",
                    "valid_to must be on or after valid_from",
                    "valid_to",
                    valid_to.isoformat(),
                )
            )
        if errors:
            return None, errors
        assert company_code is not None
        assert company_name is not None
        assert valid_from is not None
        assert tax_rate is not None
        assert deferred_tax_rate is not None
        assert loss is not None
        assert burden is not None
        return (
            TaxMasterRow(
                row_number=row_number,
                company_code=company_code,
                company_name=company_name,
                valid_from=valid_from,
                valid_to=valid_to,
                tax_rate=tax_rate,
                deferred_tax_rate=deferred_tax_rate,
                loss_carryforward=loss,
                three_year_average_tax_burden=burden,
            ),
            [],
        )


def _preflight_xlsx(payload: bytes, limits: XlsxResourceLimits) -> None:
    try:
        with ZipFile(BytesIO(payload), "r") as archive:
            members = archive.infolist()
            if len(members) > limits.max_zip_members:
                raise _resource_error("XLSX contains too many ZIP members")

            total_size = 0
            for member in members:
                total_size += member.file_size
                if member.file_size > limits.max_member_uncompressed_bytes:
                    raise _resource_error("XLSX member exceeds the uncompressed size limit")
                if total_size > limits.max_total_uncompressed_bytes:
                    raise _resource_error("XLSX exceeds the total uncompressed size limit")
                if (
                    member.file_size > 0
                    and member.file_size
                    > limits.max_compression_ratio * max(member.compress_size, 1)
                ):
                    raise _resource_error("XLSX member exceeds the compression-ratio limit")

            member_names = [member.filename for member in members]
            if len(set(member_names)) != len(member_names):
                raise _invalid_xlsx()
            worksheet_parts = _worksheet_part_names(archive, set(member_names))
            for worksheet_part in worksheet_parts:
                with archive.open(worksheet_part, "r") as worksheet_stream:
                    _inspect_worksheet_xml(worksheet_stream, limits)
    except TaxMasterWorkbookError:
        raise
    except (BadZipFile, KeyError, ParseError, RuntimeError, OSError, ValueError) as error:
        raise _invalid_xlsx() from error


def preflight_xlsx(payload: bytes, limits: XlsxResourceLimits) -> None:
    """Apply the shared XLSX archive and worksheet resource budgets."""

    if not isinstance(payload, bytes):
        raise TypeError("payload must be bytes")
    if not isinstance(limits, XlsxResourceLimits):
        raise TypeError("limits must be XlsxResourceLimits")
    _preflight_xlsx(payload, limits)


def _worksheet_part_names(archive: ZipFile, member_names: set[str]) -> tuple[str, ...]:
    content_types = _read_package_xml(archive, _CONTENT_TYPES_MEMBER, "Types")
    workbook = _read_package_xml(archive, _WORKBOOK_MEMBER, "workbook")
    relationships = _read_package_xml(
        archive,
        _WORKBOOK_RELATIONSHIPS_MEMBER,
        "Relationships",
    )

    sheet_relationship_ids = {
        relationship_id
        for element in workbook.iter()
        if _local_name(element.tag) == "sheet"
        if (relationship_id := _attribute(element, "id")) is not None
    }
    relationship_ids: set[str] = set()
    relationship_parts: set[str] = set()
    for relationship in relationships:
        if _local_name(relationship.tag) != "Relationship":
            continue
        relationship_id = relationship.attrib.get("Id")
        relationship_type = relationship.attrib.get("Type")
        target = relationship.attrib.get("Target")
        if not relationship_id or not relationship_type or not target:
            raise _invalid_xlsx()
        if relationship_id in relationship_ids:
            raise _invalid_xlsx()
        relationship_ids.add(relationship_id)

        is_worksheet = relationship_type == _WORKSHEET_RELATIONSHIP_TYPE
        is_workbook_sheet = relationship_id in sheet_relationship_ids
        if not is_workbook_sheet:
            continue
        target_mode = relationship.attrib.get("TargetMode")
        if target_mode not in (None, "Internal"):
            raise _invalid_xlsx()
        if is_workbook_sheet and relationship_type == _CHARTSHEET_RELATIONSHIP_TYPE:
            continue
        if not is_worksheet:
            raise _invalid_xlsx()
        relationship_parts.add(_normalize_relationship_target(target))

    if not sheet_relationship_ids.issubset(relationship_ids):
        raise _invalid_xlsx()

    override_parts: set[str] = set()
    content_type_parts: set[str] = set()
    for override in content_types:
        if _local_name(override.tag) != "Override":
            continue
        part_name = override.attrib.get("PartName")
        content_type = override.attrib.get("ContentType")
        if not part_name or not content_type:
            raise _invalid_xlsx()
        normalized_part = _normalize_override_part_name(part_name)
        if normalized_part in override_parts:
            raise _invalid_xlsx()
        override_parts.add(normalized_part)
        if content_type == _WORKSHEET_CONTENT_TYPE:
            content_type_parts.add(normalized_part)

    if not relationship_parts or not relationship_parts.issubset(content_type_parts):
        raise _invalid_xlsx()
    if not relationship_parts.issubset(member_names):
        raise _invalid_xlsx()
    return tuple(sorted(relationship_parts))


def _read_package_xml(archive: ZipFile, member_name: str, root_name: str) -> Element:
    with archive.open(member_name, "r") as stream:
        root = fromstring(stream.read())
    if _local_name(root.tag) != root_name:
        raise _invalid_xlsx()
    return root


def _normalize_relationship_target(target: str) -> str:
    parsed = urlsplit(target)
    if parsed.scheme or parsed.netloc or parsed.query or parsed.fragment or not parsed.path:
        raise _invalid_xlsx()
    base_parts: tuple[str, ...] = () if parsed.path.startswith("/") else ("xl",)
    return _normalize_part_path(parsed.path, base_parts)


def _normalize_override_part_name(part_name: str) -> str:
    parsed = urlsplit(part_name)
    if (
        not parsed.path.startswith("/")
        or parsed.scheme
        or parsed.netloc
        or parsed.query
        or parsed.fragment
    ):
        raise _invalid_xlsx()
    return _normalize_part_path(parsed.path, ())


def _normalize_part_path(path: str, base_parts: tuple[str, ...]) -> str:
    if "\\" in path or "\x00" in path:
        raise _invalid_xlsx()
    normalized_parts = list(base_parts)
    for part in path.split("/"):
        if part in ("", "."):
            continue
        if part == "..":
            if not normalized_parts:
                raise _invalid_xlsx()
            normalized_parts.pop()
        else:
            normalized_parts.append(part)
    if not normalized_parts:
        raise _invalid_xlsx()
    return "/".join(normalized_parts)


def _attribute(element: Element, local_name: str) -> str | None:
    values = [
        value
        for qualified_name, value in element.attrib.items()
        if _local_name(qualified_name) == local_name
    ]
    if len(values) > 1:
        raise _invalid_xlsx()
    return values[0] if values else None


def _local_name(qualified_name: str) -> str:
    return qualified_name.rsplit("}", 1)[-1]


def _inspect_worksheet_xml(stream: IO[bytes], limits: XlsxResourceLimits) -> None:
    row_count = 0
    cell_count = 0
    actual_max_row = 0
    actual_max_column = 0
    root_seen = False
    for event, element in iterparse(stream, events=("start", "end")):
        local_name = _local_name(element.tag)
        if not root_seen:
            root_seen = True
            if event != "start" or local_name != "worksheet":
                raise _invalid_xlsx()
        if event == "start" and local_name == "dimension":
            reference = element.attrib.get("ref")
            if reference is not None:
                maximum_row, dimension_cells = _dimension_extent(reference)
                if maximum_row > limits.max_worksheet_rows:
                    raise _resource_error("worksheet dimension exceeds the row limit")
                if dimension_cells > limits.max_worksheet_cells:
                    raise _resource_error("worksheet dimension exceeds the cell limit")
        elif event == "start" and local_name == "row":
            row_count += 1
            row_reference = element.attrib.get("r")
            if row_count > min(limits.max_worksheet_rows, _EXCEL_MAX_ROW):
                raise _resource_error("worksheet exceeds the row limit")
            if row_reference is not None:
                row_index = _row_number(row_reference, "worksheet row reference")
                if row_index > limits.max_worksheet_rows:
                    raise _resource_error("worksheet exceeds the row limit")
                actual_max_row = max(actual_max_row, row_index)
        elif event == "start" and local_name == "c":
            cell_count += 1
            if cell_count > limits.max_worksheet_cells:
                raise _resource_error("worksheet exceeds the cell limit")
            cell_reference = element.attrib.get("r")
            if cell_reference is None:
                raise _resource_error("worksheet cell reference is missing")
            column_index, row_index = _cell_coordinates(
                cell_reference,
                "worksheet cell reference",
            )
            if row_index > limits.max_worksheet_rows:
                raise _resource_error("worksheet cell reference exceeds the row limit")
            actual_max_row = max(actual_max_row, row_index)
            actual_max_column = max(actual_max_column, column_index)
        if event == "end":
            element.clear()
    if not root_seen:
        raise _invalid_xlsx()
    allocated_rows = max(actual_max_row, row_count)
    if allocated_rows * actual_max_column > limits.max_worksheet_cells:
        raise _resource_error("worksheet allocated rectangle exceeds the cell limit")


def _dimension_extent(reference: str) -> tuple[int, int]:
    endpoints = reference.split(":", maxsplit=1)
    start = _cell_coordinates(endpoints[0], "worksheet dimension reference")
    end = _cell_coordinates(endpoints[-1], "worksheet dimension reference")
    start_column, start_row = start
    end_column, end_row = end
    if end_column < start_column or end_row < start_row:
        raise _resource_error("worksheet dimension range is invalid")
    return end_row, (end_column - start_column + 1) * (end_row - start_row + 1)


def _cell_coordinates(reference: str, description: str) -> tuple[int, int]:
    if not 2 <= len(reference) <= _MAX_CELL_REFERENCE_LENGTH:
        raise _resource_error(f"{description} is not canonical")
    matched = _BOUNDED_CELL_REFERENCE.fullmatch(reference)
    if matched is None:
        raise _resource_error(f"{description} is not canonical")
    column = _column_number(matched.group(1))
    if column > _EXCEL_MAX_COLUMN:
        raise _resource_error(f"{description} exceeds the Excel column limit")
    return column, _row_number(matched.group(2), description)


def _row_number(reference: str, description: str) -> int:
    if not 1 <= len(reference) <= _MAX_ROW_REFERENCE_LENGTH:
        raise _resource_error(f"{description} is not canonical")
    if _BOUNDED_ROW_REFERENCE.fullmatch(reference) is None:
        raise _resource_error(f"{description} is not canonical")
    row = int(reference)
    if row > _EXCEL_MAX_ROW:
        raise _resource_error(f"{description} exceeds the Excel row limit")
    return row


def _column_number(reference: str) -> int:
    column = 0
    for character in reference:
        column = column * 26 + ord(character) - ord("A") + 1
    return column


def _resource_error(message: str) -> TaxMasterWorkbookError:
    return _workbook_error(1, "XLSX_RESOURCE_LIMIT_EXCEEDED", message)


def _invalid_xlsx() -> TaxMasterWorkbookError:
    return _workbook_error(
        1,
        "INVALID_XLSX",
        "file is not a readable XLSX workbook",
    )


def _all_blank(values: tuple[object, ...]) -> bool:
    return not values or all(value is None or value == "" for value in values)


def _formula_errors(
    row_number: int,
    value_cells: tuple[_Cell, ...],
    formula_cells: dict[int, object],
) -> list[TaxMasterRowError]:
    errors: list[TaxMasterRowError] = []
    for column_number, formula_value in sorted(formula_cells.items()):
        index = column_number - 1
        cached_value = value_cells[index].value if index < len(value_cells) else None
        if cached_value is None:
            field = REQUIRED_COLUMNS[index] if index < len(REQUIRED_COLUMNS) else None
            errors.append(
                TaxMasterRowError(
                    row_number,
                    "FORMULA_WITHOUT_CACHED_VALUE",
                    "formula cells must contain a cached value",
                    field,
                    str(formula_value),
                )
            )
    return errors


def _parse_text(
    value: object,
    row_number: int,
    field: str,
    maximum_length: int,
    errors: list[TaxMasterRowError],
) -> str | None:
    if not isinstance(value, str) or not value.strip():
        errors.append(
            TaxMasterRowError(
                row_number,
                "MISSING_VALUE",
                f"{field} is required",
                field,
                None if value is None else str(value),
            )
        )
        return None
    normalized = value.strip()
    if len(normalized) > maximum_length:
        errors.append(
            TaxMasterRowError(
                row_number,
                "VALUE_TOO_LONG",
                f"{field} must not exceed {maximum_length} characters",
                field,
                normalized,
            )
        )
        return None
    return normalized


def _parse_date(
    value: object,
    row_number: int,
    field: str,
    optional: bool,
    errors: list[TaxMasterRowError],
) -> date | None:
    if value is None or value == "":
        if optional:
            return None
        errors.append(
            TaxMasterRowError(row_number, "MISSING_VALUE", f"{field} is required", field)
        )
        return None
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif type(value) is date:
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = date.fromisoformat(value.strip())
        except ValueError:
            parsed = None
    if parsed is None:
        errors.append(
            TaxMasterRowError(
                row_number,
                "INVALID_DATE",
                f"{field} must be an Excel date or YYYY-MM-DD",
                field,
                str(value),
            )
        )
    return parsed


def _parse_rate(
    value: object,
    row_number: int,
    field: str,
    errors: list[TaxMasterRowError],
) -> Rate | None:
    try:
        if isinstance(value, bool) or value is None:
            raise ValueError
        if isinstance(value, str):
            source = value.strip()
            if source.endswith("%"):
                fraction = Decimal(source[:-1].strip()).scaleb(-2)
            else:
                fraction = Decimal(source)
        elif isinstance(value, Decimal):
            fraction = value
        elif type(value) in {int, float}:
            fraction = _excel_numeric_decimal(value)
        else:
            raise ValueError
        normalized_exponent = cast(int, fraction.normalize().as_tuple().exponent)
        if normalized_exponent < -12:
            raise ValueError
        return Rate.from_fraction(fraction)
    except _ExcelNumericPrecisionError:
        errors.append(_excel_precision_error(row_number, field, value))
        return None
    except (InvalidOperation, TypeError, ValueError):
        errors.append(
            TaxMasterRowError(
                row_number,
                "INVALID_RATE",
                f"{field} must be a finite rate between 0 and 1",
                field,
                None if value is None else str(value),
            )
        )
        return None


def _parse_loss(
    value: object,
    row_number: int,
    amount_scale: int,
    errors: list[TaxMasterRowError],
) -> Decimal | None:
    try:
        if isinstance(value, bool) or value is None:
            raise ValueError
        if isinstance(value, (Decimal, str)):
            loss = Decimal(str(value).strip())
        elif type(value) in {int, float}:
            loss = _excel_numeric_decimal(value)
        else:
            raise ValueError
        exponent = cast(int, loss.as_tuple().exponent)
        if (
            not loss.is_finite()
            or loss < 0
            or max(-exponent, 0) > amount_scale
            or not fits_database_amount(loss)
        ):
            raise ValueError
        return loss
    except _ExcelNumericPrecisionError:
        errors.append(
            _excel_precision_error(row_number, "loss_carryforward", value)
        )
        return None
    except (InvalidOperation, TypeError, ValueError):
        errors.append(
            TaxMasterRowError(
                row_number,
                "INVALID_LOSS_CARRYFORWARD",
                "loss_carryforward must be non-negative and fit the configured money scale",
                "loss_carryforward",
                None if value is None else str(value),
            )
        )
        return None


class _ExcelNumericPrecisionError(ValueError):
    pass


def _excel_numeric_decimal(value: object) -> Decimal:
    if type(value) not in {int, float}:
        raise TypeError("Excel numeric value must be int or float")
    decimal_value = Decimal(str(value))
    digits = decimal_value.as_tuple().digits
    trailing_zeros = 0
    for digit in reversed(digits):
        if digit != 0 or trailing_zeros == len(digits) - 1:
            break
        trailing_zeros += 1
    if len(digits) - trailing_zeros > 15:
        raise _ExcelNumericPrecisionError
    return decimal_value


def _excel_precision_error(
    row_number: int,
    field: str,
    value: object,
) -> TaxMasterRowError:
    return TaxMasterRowError(
        row_number,
        "EXCEL_NUMERIC_PRECISION_EXCEEDED",
        f"{field} exceeds Excel's 15-digit numeric precision; enter large values as text",
        field,
        str(value),
    )


def _period_overlap_errors(rows: list[TaxMasterRow]) -> list[TaxMasterRowError]:
    errors: list[TaxMasterRowError] = []
    seen: dict[str, list[TaxMasterRow]] = {}
    for row in rows:
        company_rows = seen.setdefault(row.company_code, [])
        for prior in company_rows:
            if row.valid_from <= (prior.valid_to or date.max) and prior.valid_from <= (
                row.valid_to or date.max
            ):
                exact_duplicate = (
                    row.valid_from == prior.valid_from and row.valid_to == prior.valid_to
                )
                errors.append(
                    TaxMasterRowError(
                        row.row_number,
                        (
                            "DUPLICATE_EFFECTIVE_PERIOD"
                            if exact_duplicate
                            else "OVERLAPPING_EFFECTIVE_PERIOD"
                        ),
                        "company effective periods must not overlap within one workbook",
                        "valid_from",
                        row.valid_from.isoformat(),
                    )
                )
                break
        company_rows.append(row)
    return errors


def _sum_losses(losses: list[Decimal], amount_scale: int) -> Decimal:
    total = Money.unrounded("0", currency="XXX", scale=amount_scale)
    for loss in losses:
        total += Money.unrounded(loss, currency="XXX", scale=amount_scale)
    return total.amount


def _workbook_error(row_number: int, error_code: str, message: str) -> TaxMasterWorkbookError:
    return TaxMasterWorkbookError((TaxMasterRowError(row_number, error_code, message),))


__all__ = [
    "DEFAULT_XLSX_RESOURCE_LIMITS",
    "REQUIRED_COLUMNS",
    "TaxMasterRow",
    "TaxMasterRowError",
    "TaxMasterWorkbookError",
    "TaxMasterXlsxAdapter",
    "XlsxResourceLimits",
    "preflight_xlsx",
]
