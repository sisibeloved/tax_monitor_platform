from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook  # type: ignore[import-untyped]

from tax_risk.application.tax_adjustment_accounts.contracts import (
    AccountCheckResult,
    AdjustmentLabel,
    AdjustmentSubject,
    CheckStatus,
)
from tax_risk.application.tax_adjustment_accounts.rules import (
    DONATION_ACCOUNT,
    WELFARE_ACCOUNT_MAX,
    WELFARE_ACCOUNT_MIN,
)


LABEL_TEXT = {
    AdjustmentLabel.WELFARE_REASONABLE: "福利费入账合理",
    AdjustmentLabel.WELFARE_BUSINESS_ENTERTAINMENT: "业务招待费异常",
    AdjustmentLabel.WELFARE_EMPLOYEE_EDUCATION: "职工教育经费异常",
    AdjustmentLabel.WELFARE_ADVERTISING_PROMOTION: "广告宣传费异常",
    AdjustmentLabel.WELFARE_CUSTOMER_GIFT_REVIEW: "广告宣传或业务招待待复核",
    AdjustmentLabel.DONATION_REASONABLE: "公益性捐赠入账合理",
    AdjustmentLabel.DONATION_SPONSORSHIP: "赞助支出异常",
    AdjustmentLabel.DONATION_ADVERTISING_PROMOTION: "广告宣传异常",
}


def render_account_check_xlsx(result: AccountCheckResult) -> bytes:
    workbook = Workbook(write_only=True)
    summary = workbook.create_sheet("公司汇总")
    summary.append(
        (
            "检查类型",
            "公司",
            "会计年度",
            "截止月份",
            "纳税调增额",
            "是否执行明细检查",
            "可检查明细数",
            "币种",
            "源记录数",
            "期间内源记录数",
            "已检查明细数",
            "金额",
            "正常明细数",
            "正常金额",
            "异常明细数",
            "异常金额",
        )
    )
    if result.currency_summaries:
        for currency_summary in result.currency_summaries:
            summary.append(
                (
                    result.request.subject.value,
                    _safe_text(result.request.company),
                    result.request.fiscal_year,
                    result.request.through_month,
                    result.adjustment_amount,
                    "是",
                    result.eligible_detail_count,
                    _safe_text(currency_summary.currency),
                    result.source_row_count,
                    result.in_scope_source_row_count,
                    currency_summary.detail_count,
                    currency_summary.amount,
                    currency_summary.normal_count,
                    currency_summary.normal_amount,
                    currency_summary.abnormal_count,
                    currency_summary.abnormal_amount,
                )
            )
    else:
        summary.append(
            (
                result.request.subject.value,
                _safe_text(result.request.company),
                result.request.fiscal_year,
                result.request.through_month,
                result.adjustment_amount,
                "否",
                result.eligible_detail_count,
                "",
                result.source_row_count,
                result.in_scope_source_row_count,
                0,
                None,
                0,
                None,
                0,
                None,
            )
        )

    monthly = workbook.create_sheet("月度汇总")
    monthly.append(
        (
            "公司",
            "会计年度",
            "期间",
            "币种",
            "明细数",
            "金额",
            "正常明细数",
            "正常金额",
            "异常明细数",
            "异常金额",
        )
    )
    for monthly_summary in result.monthly_summaries:
        monthly.append(
            (
                _safe_text(result.request.company),
                result.request.fiscal_year,
                f"{monthly_summary.month:03d}",
                _safe_text(monthly_summary.currency),
                monthly_summary.detail_count,
                monthly_summary.amount,
                monthly_summary.normal_count,
                monthly_summary.normal_amount,
                monthly_summary.abnormal_count,
                monthly_summary.abnormal_amount,
            )
        )

    details = workbook.create_sheet("全部明细")
    details.append(
        (
            "company",
            "fiscal_year",
            "fiscal_period",
            "voucher_no",
            "header_text",
            "detail_text",
            "amount_ksl",
            "gl_account",
            "account_name",
            "project_code",
            "project_name",
            "debit_credit_flag",
            "group_currency",
            "original_system_doc_no",
            "check_status",
            "check_labels",
            "matched_keywords",
        )
    )
    for detail in result.details:
        row = detail.row
        details.append(
            (
                _safe_text(row.company),
                row.fiscal_year,
                row.fiscal_period,
                _safe_text(row.voucher_no),
                _safe_text(row.header_text),
                _safe_text(row.detail_text),
                row.amount_ksl,
                _safe_text(row.gl_account),
                _safe_text(row.account_name),
                _safe_text(row.project_code),
                _safe_text(row.project_name),
                _safe_text(row.debit_credit_flag),
                _safe_text(row.group_currency),
                _safe_text(row.original_system_doc_no),
                "正常" if detail.status is CheckStatus.NORMAL else "异常",
                "、".join(LABEL_TEXT[label] for label in detail.labels),
                "、".join(detail.matched_keywords),
            )
        )

    rules = workbook.create_sheet("检查口径")
    rules.append(("项目", "口径"))
    rules.append(("金额字段", "amount_ksl；group_currency 仅作为币种"))
    rules.append(("期间", f"001-{result.request.through_month:03d}"))
    rules.append(("明细检查门控", "纳税调增额严格大于0时才执行；否则保留汇总并跳过明细标签"))
    if result.request.subject is AdjustmentSubject.WELFARE:
        rules.append(("科目范围", f"{WELFARE_ACCOUNT_MIN}-{WELFARE_ACCOUNT_MAX}（含边界）"))
        rules.append(("业务招待", "客户、供应商、政府接待、商务宴请"))
        rules.append(("职工教育", "培训费、讲师费、考试费"))
        rules.append(("广告宣传", "宣传赠品"))
        rules.append(("待复核", "客户礼品"))
    else:
        rules.append(("科目", DONATION_ACCOUNT))
        rules.append(("赞助", "赞助"))
        rules.append(("广告宣传", "冠名、广告权益、品牌露出"))

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _safe_text(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


__all__ = ["LABEL_TEXT", "render_account_check_xlsx"]
