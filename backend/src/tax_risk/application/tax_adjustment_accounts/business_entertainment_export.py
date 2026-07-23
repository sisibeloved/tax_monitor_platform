from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook  # type: ignore[import-untyped]

from tax_risk.application.tax_adjustment_accounts.business_entertainment import (
    BUSINESS_ENTERTAINMENT_ACCOUNT,
    EMPLOYEE_WELFARE_KEYWORDS,
    MEETING_OR_EDUCATION_KEYWORDS,
    BusinessEntertainmentCheckResult,
    BusinessEntertainmentLabel,
)
from tax_risk.application.tax_adjustment_accounts.contracts import CheckStatus


LABEL_TEXT = {
    BusinessEntertainmentLabel.REASONABLE: "业务招待费入账合理",
    BusinessEntertainmentLabel.EMPLOYEE_WELFARE: "可能应归福利费",
    BusinessEntertainmentLabel.MEETING_OR_EDUCATION: "可能应归会议费或职工教育经费",
}


def render_business_entertainment_check_xlsx(
    result: BusinessEntertainmentCheckResult,
) -> bytes:
    workbook = Workbook(write_only=True)
    summary = workbook.create_sheet("公司汇总")
    summary.append(
        (
            "公司",
            "会计年度",
            "截止月份",
            "币种",
            "源记录数",
            "期间内源记录数",
            "业务招待费明细数",
            "金额",
            "正常明细数",
            "正常金额",
            "异常明细数",
            "异常金额",
            "和思明细源记录数",
            "和思发票源记录数",
            "申请单源记录数",
        )
    )
    if result.currency_summaries:
        for currency_summary in result.currency_summaries:
            summary.append(
                (
                    _safe_text(result.request.company),
                    result.request.fiscal_year,
                    result.request.through_month,
                    _safe_text(currency_summary.currency),
                    result.source_row_count,
                    result.in_scope_source_row_count,
                    currency_summary.detail_count,
                    currency_summary.amount,
                    currency_summary.normal_count,
                    currency_summary.normal_amount,
                    currency_summary.abnormal_count,
                    currency_summary.abnormal_amount,
                    result.hesi_detail_source_row_count,
                    result.hesi_invoice_source_row_count,
                    result.hesi_application_source_row_count,
                )
            )
    else:
        summary.append(
            (
                _safe_text(result.request.company),
                result.request.fiscal_year,
                result.request.through_month,
                "",
                result.source_row_count,
                result.in_scope_source_row_count,
                0,
                None,
                0,
                None,
                0,
                None,
                result.hesi_detail_source_row_count,
                result.hesi_invoice_source_row_count,
                result.hesi_application_source_row_count,
            )
        )

    monthly = workbook.create_sheet("月度汇总")
    monthly.append(
        (
            "公司",
            "会计年度",
            "期间",
            "币种",
            "明细数",
            "金额",
            "正常明细数",
            "正常金额",
            "异常明细数",
            "异常金额",
        )
    )
    for monthly_summary in result.monthly_summaries:
        monthly.append(
            (
                _safe_text(result.request.company),
                result.request.fiscal_year,
                f"{monthly_summary.month:03d}",
                _safe_text(monthly_summary.currency),
                monthly_summary.detail_count,
                monthly_summary.amount,
                monthly_summary.normal_count,
                monthly_summary.normal_amount,
                monthly_summary.abnormal_count,
                monthly_summary.abnormal_amount,
            )
        )

    details = workbook.create_sheet("全部明细")
    details.append(
        (
            "company",
            "fiscal_year",
            "fiscal_period",
            "voucher_no",
            "header_text",
            "detail_text",
            "amount_ksl",
            "gl_account",
            "account_name",
            "project_code",
            "project_name",
            "debit_credit_flag",
            "group_currency",
            "original_system_doc_no",
            "check_status",
            "check_labels",
            "matched_keywords",
            "decision_source",
            "evaluated_sources",
            "evidence_texts",
            "hesi_document_code",
            "hesi_detail_match_count",
            "hesi_invoice_match_count",
            "reception_apply_codes",
            "hesi_application_match_count",
        )
    )
    for detail in result.details:
        row = detail.row
        details.append(
            (
                _safe_text(row.company),
                row.fiscal_year,
                row.fiscal_period,
                _safe_text(row.voucher_no),
                _safe_text(row.header_text),
                _safe_text(row.detail_text),
                row.amount_ksl,
                _safe_text(row.gl_account),
                _safe_text(row.account_name),
                _safe_text(row.project_code),
                _safe_text(row.project_name),
                _safe_text(row.debit_credit_flag),
                _safe_text(row.group_currency),
                _safe_text(row.original_system_doc_no),
                "正常" if detail.status is CheckStatus.NORMAL else "异常",
                _safe_text("、".join(LABEL_TEXT[label] for label in detail.labels)),
                _safe_text("、".join(detail.matched_keywords)),
                detail.decision_source.value,
                "、".join(source.value for source in detail.evaluated_sources),
                _safe_text(" | ".join(detail.evidence_texts)),
                _safe_text(detail.hesi_document_code or ""),
                detail.hesi_detail_match_count,
                detail.hesi_invoice_match_count,
                _safe_text("、".join(detail.reception_apply_codes)),
                detail.hesi_application_match_count,
            )
        )

    rules = workbook.create_sheet("检查口径")
    rules.append(("项目", "口径"))
    rules.append(("科目", BUSINESS_ENTERTAINMENT_ACCOUNT))
    rules.append(("金额字段", "amount_ksl；按group_currency分别汇总"))
    rules.append(("期间", f"001-{result.request.through_month:03d}（本年累计）"))
    rules.append(("福利费候选", "、".join(EMPLOYEE_WELFARE_KEYWORDS)))
    rules.append(("会议费或职工教育候选", "、".join(MEETING_OR_EDUCATION_KEYWORDS)))
    rules.append(
        (
            "证据链",
            "settlement_adjustment.detail_text -> HS单号去掉前缀 -> "
            "hesimingxi.description -> hesiinvoice.reception_apply_code -> "
            "apply.description",
        )
    )
    rules.append(("正常判断", "证据链中均未命中上述关键词时，认定业务招待费入账合理"))

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _safe_text(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


__all__ = ["LABEL_TEXT", "render_business_entertainment_check_xlsx"]
