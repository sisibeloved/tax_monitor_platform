"""Regenerate the committed tax-master workbook fixtures with openpyxl."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook


HEADERS = (
    "company_code",
    "company_name",
    "valid_from",
    "valid_to",
    "tax_rate",
    "loss_carryforward",
    "three_year_average_tax_burden",
)
FIXTURE_DIR = Path(__file__).resolve().parent


def _workbook(rows: list[tuple[object, ...]], destination: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "tax_master"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row_number, 3).number_format = "yyyy-mm-dd"
        worksheet.cell(row_number, 4).number_format = "yyyy-mm-dd"
    worksheet["E3"].number_format = "0.00%"
    workbook.save(destination)
    workbook.close()


def main() -> None:
    _workbook(
        [
            (
                "C001",
                "Company One",
                date(2026, 1, 1),
                None,
                "25%",
                "100000.00",
                "9%",
            ),
            (
                "C002",
                "Company Two",
                date(2026, 4, 1),
                date(2026, 12, 31),
                0.25,
                0,
                0.08,
            ),
        ],
        FIXTURE_DIR / "tax_master_valid.xlsx",
    )
    _workbook(
        [
            (
                "C001",
                "Company One",
                date(2026, 1, 1),
                date(2026, 6, 30),
                "25%",
                0,
                "9%",
            ),
            (
                "C001",
                "Company One",
                date(2026, 6, 1),
                None,
                "25%",
                0,
                "9%",
            ),
        ],
        FIXTURE_DIR / "tax_master_duplicate.xlsx",
    )


if __name__ == "__main__":
    main()
