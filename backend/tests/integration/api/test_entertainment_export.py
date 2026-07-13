from __future__ import annotations

from datetime import date
from functools import partial
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.integration.application.test_semantic_case_routing import (
    _detection,
    _seed_graph,
)
from tax_risk.application.business_entertainment.export import escape_excel_text
from tax_risk.application.exports import InMemoryExportObjectStore
from tax_risk.application.case_merge import CaseMergeService
from tax_risk.application.semantic.detection_router import SemanticCaseRouter
from tax_risk.config import Settings
from tax_risk.main import create_app
from tax_risk.persistence.repositories import UnitOfWork, create_session_factory
from tax_risk.security.principal import COMPANY_FINANCE_ROLE, Principal
from tax_risk.security.service_scope import issue_service_scope_token
from tax_risk.workers.celery_app import create_celery_app
from tax_risk.workers.exports import RENDER_EXPORT_TASK, register_export_tasks


def test_export_contains_one_safe_root_case_row(
    isolated_database_url: str,
) -> None:
    engine, factory = create_session_factory(isolated_database_url)
    try:
        company_code, company_id, snapshot_id, source_id, _, link_id = _seed_graph(engine)
        routed = SemanticCaseRouter(partial(UnitOfWork, factory)).route(
            _detection(
                company_code=company_code,
                snapshot_id=snapshot_id,
                source_id=source_id,
                label="MEETING_EXPENSE",
                model_version="export-v1",
            ),
            suspicious_labels=frozenset({"MEETING_EXPENSE"}),
        )
        assert routed.risk_case_id is not None
        CaseMergeService(partial(UnitOfWork, factory)).resolve_to_sap(
            business_case_id=routed.risk_case_id,
            evidence_link_id=link_id,
            expected_row_version=1,
            actor="reviewer",
        )
        principal = Principal(
            subject="group-tax@example.com",
            roles=frozenset({COMPANY_FINANCE_ROLE}),
            allowed_company_ids=frozenset({company_id}),
            organization_path="/companies/scoped",
        )
        settings = Settings(
            environment="test",
            export_download_secret="test-secret",
            worker_scope_secret="signed-export-api-scope-test",
            redis_url="redis://localhost:6379/15",
            celery_task_always_eager=True,
            celery_task_eager_propagates=True,
            celery_task_store_eager_result=True,
        )
        dispatched: list[dict[str, object]] = []
        app = create_app(
            uow_factory=partial(UnitOfWork, factory),
            settings=settings,
            principal_provider=lambda _request: principal,
            export_dispatcher=lambda **payload: dispatched.append(payload),
            export_object_store=InMemoryExportObjectStore(),
        )
        with TestClient(app) as client:
            created = client.post(
                "/api/v1/exports",
                json={"export_type": "BUSINESS_ENTERTAINMENT", "filters": {}},
            )
            assert created.status_code == 202, created.text
            job_id = created.json()["id"]
            worker = create_celery_app(settings)
            register_export_tasks(
                app=worker,
                service_factory=lambda: app.state.export_service,
            )
            scope_token = issue_service_scope_token(
                secret=settings.worker_scope_secret,
                queue="exports",
                run_type="EXPORT",
                batch_id=job_id,
                company_ids=frozenset({company_id}),
                period=date(2026, 6, 30),
            )
            worker.signature(
                RENDER_EXPORT_TASK,
                args=(
                    job_id,
                    dispatched[0]["authorization_version"],
                    scope_token,
                ),
            ).apply_async().get(timeout=10)
            issued = client.post(f"/api/v1/exports/{job_id}/download-url")
            assert issued.status_code == 200, issued.text
            response = client.get(issued.json()["url"])

        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        assert len(rows) == 2
        assert rows[1][1] == company_code
        assert rows[1][4] == "SAP_LINKED"
        assert rows[1][8] == 100
        assert escape_excel_text("=1+1") == "'=1+1"
    finally:
        engine.dispose()
