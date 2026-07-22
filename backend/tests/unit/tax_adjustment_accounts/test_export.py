from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from tax_risk.application.tax_adjustment_accounts.contracts import (
    AccountCheckRequest,
    AdjustmentSubject,
    SettlementAdjustmentRow,
)
from tax_risk.application.tax_adjustment_accounts.export import render_account_check_xlsx
from tax_risk.application.tax_adjustment_accounts.service import (
    TaxAdjustmentAccountCheckService,
)


class Source:
    def fetch_rows(
        self,
        *,
        company: str,
        fiscal_year: str,
    ) -> tuple[SettlementAdjustmentRow, ...]:
        return (
            _row(company, fiscal_year, "1", "公益捐赠", "100"),
            _row(company, fiscal_year, "2", "=赞助公式", "20"),
        )


def _row(
    company: str,
    fiscal_year: str,
    voucher: str,
    detail: str,
    amount: str,
) -> SettlementAdjustmentRow:
    return SettlementAdjustmentRow(
        company=company,
        fiscal_year=fiscal_year,
        fiscal_period="001",
        voucher_no=voucher,
        header_text="",
        detail_text=detail,
        amount_ksl=Decimal(amount),
        gl_account="6711060000",
        account_name="公益性捐赠",
        project_code="",
        project_name="",
        debit_credit_flag="S",
        group_currency="CNY",
        original_system_doc_no=f"source-{voucher}",
    )


def _request() -> AccountCheckRequest:
    return AccountCheckRequest(
        subject=AdjustmentSubject.DONATION,
        company="3320",
        fiscal_year="2025",
        through_month=1,
    )


def test_export_contains_selected_details_and_escapes_formulas() -> None:
    result = TaxAdjustmentAccountCheckService(Source()).check(
        _request(),
        adjustment_amount=Decimal("1"),
    )

    workbook = load_workbook(BytesIO(render_account_check_xlsx(result)), data_only=False)
    try:
        summary = workbook["公司汇总"]
        summary_headers = [cell.value for cell in summary[1]]
        summary_row = dict(
            zip(summary_headers, next(summary.iter_rows(min_row=2, values_only=True)))
        )
        worksheet = workbook["全部明细"]
        headers = [cell.value for cell in worksheet[1]]
        rows = [
            dict(zip(headers, values))
            for values in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()

    assert summary_row["是否执行明细检查"] == "是"
    assert len(rows) == 2
    assert {row["check_status"] for row in rows} == {"正常", "异常"}
    assert rows[1]["detail_text"] == "'=赞助公式"
    assert rows[1]["amount_ksl"] == 20


def test_export_keeps_summary_but_omits_details_when_gate_is_closed() -> None:
    result = TaxAdjustmentAccountCheckService(Source()).check_rows(
        _request(),
        source_rows=Source().fetch_rows(company="3320", fiscal_year="2025"),
        adjustment_amount=Decimal("0"),
    )

    workbook = load_workbook(BytesIO(render_account_check_xlsx(result)), data_only=False)
    try:
        summary = workbook["公司汇总"]
        headers = [cell.value for cell in summary[1]]
        row = dict(zip(headers, next(summary.iter_rows(min_row=2, values_only=True))))
        details = workbook["全部明细"]
        detail_rows = list(details.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert row["纳税调增额"] == 0
    assert row["是否执行明细检查"] == "否"
    assert row["可检查明细数"] == 2
    assert detail_rows == []
