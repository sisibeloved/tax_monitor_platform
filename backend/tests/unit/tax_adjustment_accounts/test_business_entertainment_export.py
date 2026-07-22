from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from tax_risk.application.tax_adjustment_accounts.business_entertainment import (
    BusinessEntertainmentAccountCheckService,
    BusinessEntertainmentCheckRequest,
    HesiApplicationRow,
    HesiDetailRow,
    HesiInvoiceRow,
)
from tax_risk.application.tax_adjustment_accounts.business_entertainment_export import (
    render_business_entertainment_check_xlsx,
)
from tax_risk.application.tax_adjustment_accounts.contracts import SettlementAdjustmentRow


class Source:
    def __init__(self, rows: tuple[object, ...]) -> None:
        self.rows = rows

    def fetch_rows(self, **_: str) -> tuple[object, ...]:
        return self.rows


def test_export_contains_evidence_chain_and_escapes_external_formulas() -> None:
    settlement_rows = (
        SettlementAdjustmentRow(
            company="3HD0",
            fiscal_year="2025",
            fiscal_period="001",
            voucher_no="1",
            detail_text="客户接待",
            amount_ksl=Decimal("100"),
            gl_account="6600400000",
            account_name="业务招待费",
            group_currency="CNY",
            original_system_doc_no="HSB1",
        ),
    )
    result = BusinessEntertainmentAccountCheckService(
        settlement_source=Source(settlement_rows),  # type: ignore[arg-type]
        hesi_detail_source=Source(
            (HesiDetailRow(company_code="3HD0", document_code="B1", description="接待"),)
        ),  # type: ignore[arg-type]
        hesi_invoice_source=Source(
            (HesiInvoiceRow(company_code="3HD0", code="B1", reception_apply_code="A1"),)
        ),  # type: ignore[arg-type]
        hesi_application_source=Source(
            (HesiApplicationRow(company_code="3HD0", code="A1", description="=会议通知"),)
        ),  # type: ignore[arg-type]
    ).run(
        BusinessEntertainmentCheckRequest(
            company="3HD0",
            fiscal_year="2025",
            through_month=1,
        )
    )

    workbook = load_workbook(
        BytesIO(render_business_entertainment_check_xlsx(result)),
        data_only=False,
    )
    try:
        summary_headers = [cell.value for cell in workbook["公司汇总"][1]]
        summary = dict(
            zip(
                summary_headers,
                next(workbook["公司汇总"].iter_rows(min_row=2, values_only=True)),
            )
        )
        detail_headers = [cell.value for cell in workbook["全部明细"][1]]
        detail = dict(
            zip(
                detail_headers,
                next(workbook["全部明细"].iter_rows(min_row=2, values_only=True)),
            )
        )
        rule_rows = list(workbook["检查口径"].iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert summary["金额"] == 100
    assert summary["申请单源记录数"] == 1
    assert detail["check_status"] == "异常"
    assert detail["hesi_document_code"] == "B1"
    assert detail["reception_apply_codes"] == "A1"
    assert detail["evidence_texts"] == "'=会议通知"
    assert any(name == "证据链" and "hesiinvoice" in value for name, value in rule_rows)
