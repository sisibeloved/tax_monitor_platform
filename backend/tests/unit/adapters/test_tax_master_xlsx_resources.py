from __future__ import annotations

from dataclasses import replace
from datetime import date
from io import BytesIO
from math import ceil
from time import perf_counter
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from tax_risk.adapters.ingest import tax_master_xlsx as xlsx_module
from tax_risk.adapters.ingest.tax_master_xlsx import (
    DEFAULT_XLSX_RESOURCE_LIMITS,
    TaxMasterWorkbookError,
    TaxMasterXlsxAdapter,
    XlsxResourceLimits,
)


HEADERS = (
    "company_code",
    "company_name",
    "valid_from",
    "valid_to",
    "tax_rate",
    "deferred_tax_rate",
    "loss_carryforward",
    "three_year_average_tax_burden",
)


def _xlsx(row_count: int = 1) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "tax_master"
    worksheet.append(HEADERS)
    for index in range(row_count):
        worksheet.append(
            (
                f"C{index:04d}",
                f"Company {index}",
                date(2026, 1, 1),
                None,
                "25%",
                "20%",
                "100.00",
                "9%",
            )
        )
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _rewrite_zip(
    payload: bytes,
    *,
    replacements: dict[str, bytes] | None = None,
    added_members: dict[str, bytes] | None = None,
    removed_members: set[str] | None = None,
    extra_members: int = 0,
) -> bytes:
    output = BytesIO()
    replacements = replacements or {}
    added_members = added_members or {}
    removed_members = removed_members or set()
    with ZipFile(BytesIO(payload), "r") as source, ZipFile(
        output, "w", compression=ZIP_DEFLATED
    ) as target:
        for info in source.infolist():
            if info.filename in removed_members:
                continue
            target.writestr(info.filename, replacements.get(info.filename, source.read(info)))
        for name, content in added_members.items():
            target.writestr(name, content)
        for index in range(extra_members):
            target.writestr(f"extra/member-{index}.txt", b"x")
    return output.getvalue()


def _assert_resource_rejected(payload: bytes, limits: XlsxResourceLimits) -> None:
    with pytest.raises(TaxMasterWorkbookError) as caught:
        TaxMasterXlsxAdapter(payload, amount_scale=2, limits=limits).parse()
    assert caught.value.errors[0].error_code == "XLSX_RESOURCE_LIMIT_EXCEEDED"


def _assert_fast_preflight_rejection(
    payload: bytes,
    limits: XlsxResourceLimits,
    monkeypatch: pytest.MonkeyPatch,
    *,
    expected_error_code: str = "XLSX_RESOURCE_LIMIT_EXCEEDED",
) -> None:
    load_calls = 0

    def forbidden_load(*args: object, **kwargs: object) -> object:
        nonlocal load_calls
        load_calls += 1
        raise AssertionError("load_workbook must not run before resource rejection")

    monkeypatch.setattr(xlsx_module, "load_workbook", forbidden_load)
    started = perf_counter()
    with pytest.raises(TaxMasterWorkbookError) as caught:
        TaxMasterXlsxAdapter(payload, amount_scale=2, limits=limits).parse()
    elapsed = perf_counter() - started

    assert caught.value.errors[0].error_code == expected_error_code
    assert load_calls == 0
    assert elapsed < 1.0


def _assert_invalid_before_openpyxl(
    payload: bytes,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _assert_fast_preflight_rejection(
        payload,
        DEFAULT_XLSX_RESOURCE_LIMITS,
        monkeypatch,
        expected_error_code="INVALID_XLSX",
    )


def _worksheet_xml(payload: bytes) -> bytes:
    with ZipFile(BytesIO(payload)) as archive:
        return archive.read("xl/worksheets/sheet1.xml")


def _with_worksheet_xml(payload: bytes, worksheet_xml: bytes) -> bytes:
    return _rewrite_zip(
        payload,
        replacements={"xl/worksheets/sheet1.xml": worksheet_xml},
    )


def _with_relocated_worksheet(
    payload: bytes,
    *,
    worksheet_xml: bytes | None = None,
    relationship_target: str = "/xl/custom-sheet.xml",
) -> bytes:
    source_member = "xl/worksheets/sheet1.xml"
    target_member = "xl/custom-sheet.xml"
    source_part = f"/{source_member}".encode()
    target_part = f"/{target_member}".encode()
    with ZipFile(BytesIO(payload), "r") as archive:
        relocated_xml = (
            archive.read(source_member) if worksheet_xml is None else worksheet_xml
        )
        relationships = archive.read("xl/_rels/workbook.xml.rels").replace(
            b'Target="' + source_part + b'"',
            f'Target="{relationship_target}"'.encode(),
        )
        content_types = archive.read("[Content_Types].xml").replace(
            b'PartName="' + source_part + b'"',
            b'PartName="' + target_part + b'"',
        )
    return _rewrite_zip(
        payload,
        replacements={
            "xl/_rels/workbook.xml.rels": relationships,
            "[Content_Types].xml": content_types,
        },
        added_members={target_member: relocated_xml},
        removed_members={source_member},
    )


def test_high_compression_ratio_is_rejected_before_openpyxl() -> None:
    payload = _xlsx()
    compressed_xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        + b" " * 200_000
        + b"<sheetData/></worksheet>"
    )
    payload = _rewrite_zip(
        payload,
        replacements={"xl/worksheets/sheet1.xml": compressed_xml},
    )
    limits = replace(
        DEFAULT_XLSX_RESOURCE_LIMITS,
        max_member_uncompressed_bytes=1_000_000,
        max_total_uncompressed_bytes=2_000_000,
        max_compression_ratio=10,
    )

    _assert_resource_rejected(payload, limits)


def test_zip_member_count_is_bounded() -> None:
    payload = _xlsx()
    with ZipFile(BytesIO(payload)) as archive:
        member_count = len(archive.infolist())
    payload = _rewrite_zip(payload, extra_members=2)
    limits = replace(DEFAULT_XLSX_RESOURCE_LIMITS, max_zip_members=member_count + 1)

    _assert_resource_rejected(payload, limits)


def test_member_and_total_uncompressed_sizes_are_bounded() -> None:
    payload = _xlsx()
    with ZipFile(BytesIO(payload)) as archive:
        infos = archive.infolist()
        maximum_member = max(info.file_size for info in infos)
        total_size = sum(info.file_size for info in infos)

    _assert_resource_rejected(
        payload,
        replace(
            DEFAULT_XLSX_RESOURCE_LIMITS,
            max_member_uncompressed_bytes=maximum_member - 1,
        ),
    )
    _assert_resource_rejected(
        payload,
        replace(
            DEFAULT_XLSX_RESOURCE_LIMITS,
            max_total_uncompressed_bytes=total_size - 1,
        ),
    )


@pytest.mark.parametrize(
    "limits",
    [
        replace(DEFAULT_XLSX_RESOURCE_LIMITS, max_worksheet_rows=2),
        replace(DEFAULT_XLSX_RESOURCE_LIMITS, max_worksheet_cells=20),
    ],
)
def test_worksheet_rows_and_cells_are_bounded(limits: XlsxResourceLimits) -> None:
    _assert_resource_rejected(_xlsx(row_count=2), limits)


def test_malicious_dimension_is_rejected_before_row_tuple_allocation() -> None:
    payload = _xlsx()
    with ZipFile(BytesIO(payload)) as archive:
        worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
    worksheet_xml = worksheet_xml.replace(
        b'<dimension ref="A1:H2" />',
        b'<dimension ref="A1:XFD1048576" />',
    )
    payload = _rewrite_zip(
        payload,
        replacements={"xl/worksheets/sheet1.xml": worksheet_xml},
    )

    _assert_resource_rejected(payload, DEFAULT_XLSX_RESOURCE_LIMITS)


def test_relocated_worksheet_limit_is_rejected_before_openpyxl(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload).replace(b'r="H2"', b'r="XFE2"')
    payload = _with_relocated_worksheet(payload, worksheet_xml=worksheet_xml)

    _assert_fast_preflight_rejection(
        payload,
        DEFAULT_XLSX_RESOURCE_LIMITS,
        monkeypatch,
    )


def test_relocated_worksheet_within_limits_is_parsed() -> None:
    payload = _with_relocated_worksheet(
        _xlsx(),
        relationship_target="custom-sheet.xml",
    )

    parsed = TaxMasterXlsxAdapter(payload, amount_scale=2).parse()

    assert [(row.row_number, row.company_code) for row in parsed] == [(2, "C0000")]


def test_worksheet_relationship_path_escape_is_rejected_before_openpyxl(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _with_relocated_worksheet(
        _xlsx(),
        relationship_target="../../xl/custom-sheet.xml",
    )

    _assert_invalid_before_openpyxl(payload, monkeypatch)


def test_external_worksheet_relationship_is_rejected_before_openpyxl(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _with_relocated_worksheet(_xlsx())
    with ZipFile(BytesIO(payload)) as archive:
        relationships = archive.read("xl/_rels/workbook.xml.rels")
    relationships = relationships.replace(
        b'Target="/xl/custom-sheet.xml"',
        b'Target="/xl/custom-sheet.xml" TargetMode="External"',
    )
    payload = _rewrite_zip(
        payload,
        replacements={"xl/_rels/workbook.xml.rels": relationships},
    )

    _assert_invalid_before_openpyxl(payload, monkeypatch)


def test_missing_relocated_worksheet_part_is_rejected_before_openpyxl(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _rewrite_zip(
        _with_relocated_worksheet(_xlsx()),
        removed_members={"xl/custom-sheet.xml"},
    )

    _assert_invalid_before_openpyxl(payload, monkeypatch)


def test_worksheet_relationship_content_type_mismatch_is_rejected_before_openpyxl(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _with_relocated_worksheet(_xlsx())
    with ZipFile(BytesIO(payload)) as archive:
        content_types = archive.read("[Content_Types].xml")
    content_types = content_types.replace(
        b"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
        b"application/xml",
    )
    payload = _rewrite_zip(
        payload,
        replacements={"[Content_Types].xml": content_types},
    )

    _assert_invalid_before_openpyxl(payload, monkeypatch)


def test_nonworksheet_xml_is_not_subject_to_worksheet_coordinate_limits() -> None:
    custom_xml = b'<custom><row r="9999999"><c r="XFE9999999" /></row></custom>'
    payload = _rewrite_zip(
        _xlsx(),
        added_members={"xl/custom-metadata.xml": custom_xml},
    )

    parsed = TaxMasterXlsxAdapter(payload, amount_scale=2).parse()

    assert [(row.row_number, row.company_code) for row in parsed] == [(2, "C0000")]


def test_stale_unreferenced_worksheet_relationship_is_ignored() -> None:
    payload = _xlsx()
    with ZipFile(BytesIO(payload), "r") as archive:
        relationships = archive.read("xl/_rels/workbook.xml.rels").replace(
            b"</Relationships>",
            (
                b'<Relationship Id="rIdStale" '
                b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships/worksheet" Target="worksheets/stale.xml"/>'
                b"</Relationships>"
            ),
        )
        content_types = archive.read("[Content_Types].xml").replace(
            b"</Types>",
            (
                b'<Override PartName="/xl/worksheets/stale.xml" '
                b'ContentType="application/vnd.openxmlformats-officedocument.'
                b'spreadsheetml.worksheet+xml"/>'
                b"</Types>"
            ),
        )
    payload = _rewrite_zip(
        payload,
        replacements={
            "xl/_rels/workbook.xml.rels": relationships,
            "[Content_Types].xml": content_types,
        },
    )

    parsed = TaxMasterXlsxAdapter(payload, amount_scale=2).parse()

    assert [(row.row_number, row.company_code) for row in parsed] == [(2, "C0000")]


@pytest.mark.parametrize("location", ["dimension", "cell"])
def test_very_long_column_reference_is_rejected_in_constant_time_before_openpyxl(
    location: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload)
    long_column = b"A" * 100_000
    if location == "dimension":
        worksheet_xml = worksheet_xml.replace(
            b'<dimension ref="A1:H2" />',
            b'<dimension ref="A1:' + long_column + b'2" />',
        )
    else:
        worksheet_xml = worksheet_xml.replace(b'r="H2"', b'r="' + long_column + b'2"')
    payload = _with_worksheet_xml(payload, worksheet_xml)
    limits = replace(
        DEFAULT_XLSX_RESOURCE_LIMITS,
        max_compression_ratio=1_000_000,
    )

    _assert_fast_preflight_rejection(payload, limits, monkeypatch)


@pytest.mark.parametrize("location", ["dimension", "row", "cell"])
def test_very_long_row_reference_is_rejected_before_decimal_conversion(
    location: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload)
    long_row = b"1" * 100_000
    if location == "dimension":
        worksheet_xml = worksheet_xml.replace(
            b'<dimension ref="A1:H2" />',
            b'<dimension ref="A1:H' + long_row + b'" />',
        )
    elif location == "row":
        worksheet_xml = worksheet_xml.replace(b'<row r="2">', b'<row r="' + long_row + b'">')
    else:
        worksheet_xml = worksheet_xml.replace(b'r="H2"', b'r="H' + long_row + b'"')
    payload = _with_worksheet_xml(payload, worksheet_xml)
    limits = replace(
        DEFAULT_XLSX_RESOURCE_LIMITS,
        max_compression_ratio=1_000_000,
    )

    _assert_fast_preflight_rejection(payload, limits, monkeypatch)


@pytest.mark.parametrize("location", ["dimension", "cell"])
@pytest.mark.parametrize("column", ["XFE", "ZZZ"])
def test_column_beyond_excel_grid_is_rejected_before_openpyxl(
    location: str,
    column: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload)
    if location == "dimension":
        worksheet_xml = worksheet_xml.replace(
            b'<dimension ref="A1:H2" />',
            f'<dimension ref="A1:{column}2" />'.encode(),
        )
    else:
        worksheet_xml = worksheet_xml.replace(b'r="H2"', f'r="{column}2"'.encode())
    payload = _with_worksheet_xml(payload, worksheet_xml)

    _assert_fast_preflight_rejection(
        payload,
        DEFAULT_XLSX_RESOURCE_LIMITS,
        monkeypatch,
    )


@pytest.mark.parametrize("location", ["dimension", "row", "cell"])
def test_row_beyond_excel_grid_is_rejected_even_when_configured_limit_is_higher(
    location: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload)
    if location == "dimension":
        worksheet_xml = worksheet_xml.replace(
            b'<dimension ref="A1:H2" />',
            b'<dimension ref="A1:A1048577" />',
        )
    elif location == "row":
        worksheet_xml = worksheet_xml.replace(b'<row r="2">', b'<row r="1048577">')
    else:
        worksheet_xml = worksheet_xml.replace(b'r="H2"', b'r="H1048577"')
    payload = _with_worksheet_xml(payload, worksheet_xml)
    limits = replace(
        DEFAULT_XLSX_RESOURCE_LIMITS,
        max_worksheet_rows=2_000_000,
        max_worksheet_cells=20_000_000,
    )

    _assert_fast_preflight_rejection(payload, limits, monkeypatch)


@pytest.mark.parametrize("location", ["dimension", "cell"])
def test_xfd_excel_column_boundary_passes_preflight(location: str) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload)
    if location == "dimension":
        worksheet_xml = worksheet_xml.replace(
            b'<dimension ref="A1:H2" />',
            b'<dimension ref="A1:XFD2" />',
        )
    else:
        worksheet_xml = worksheet_xml.replace(b'r="H2"', b'r="XFD2"')
    payload = _with_worksheet_xml(payload, worksheet_xml)
    limits = replace(
        DEFAULT_XLSX_RESOURCE_LIMITS,
        max_worksheet_rows=2,
        max_worksheet_cells=32_768,
    )

    xlsx_module._preflight_xlsx(payload, limits)


@pytest.mark.parametrize("variant", ["sparse_far_column", "far_row_and_column"])
def test_actual_sparse_coordinate_rectangle_cannot_hide_behind_low_dimension(
    variant: str,
) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload).replace(
        b'<dimension ref="A1:H2" />',
        b'<dimension ref="A1:A1" />',
    )
    if variant == "sparse_far_column":
        worksheet_xml = worksheet_xml.replace(b'r="H2"', b'r="ZZ1"')
    else:
        worksheet_xml = worksheet_xml.replace(b'<row r="2">', b'<row r="10">')
        for column in b"ABCDEFG":
            worksheet_xml = worksheet_xml.replace(
                b'r="' + bytes((column,)) + b'2"',
                b'r="' + bytes((column,)) + b'10"',
            )
        worksheet_xml = worksheet_xml.replace(b'r="G10"', b'r="Z10"')
    payload = _with_worksheet_xml(payload, worksheet_xml)
    limits = replace(
        DEFAULT_XLSX_RESOURCE_LIMITS,
        max_worksheet_rows=20,
        max_worksheet_cells=20,
    )

    _assert_resource_rejected(payload, limits)


@pytest.mark.parametrize("row_reference", ["21.0", "1e9"])
def test_noncanonical_row_reference_is_rejected_during_preflight(
    row_reference: str,
) -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload).replace(
        b'<row r="2">',
        f'<row r="{row_reference}">'.encode(),
    )
    payload = _with_worksheet_xml(payload, worksheet_xml)
    limits = replace(DEFAULT_XLSX_RESOURCE_LIMITS, max_worksheet_rows=20)

    _assert_resource_rejected(payload, limits)


def test_low_reported_dimension_with_actual_coordinates_in_budget_is_parsed() -> None:
    payload = _xlsx()
    worksheet_xml = _worksheet_xml(payload).replace(
        b'<dimension ref="A1:H2" />',
        b'<dimension ref="A1:A1" />',
    )
    payload = _with_worksheet_xml(payload, worksheet_xml)
    limits = replace(
        DEFAULT_XLSX_RESOURCE_LIMITS,
        max_worksheet_rows=2,
        max_worksheet_cells=16,
    )

    assert len(TaxMasterXlsxAdapter(payload, amount_scale=2, limits=limits).parse()) == 1


def test_package_at_all_configured_boundaries_is_accepted() -> None:
    payload = _xlsx()
    with ZipFile(BytesIO(payload)) as archive:
        infos = archive.infolist()
        maximum_ratio = max(
            ceil(info.file_size / max(info.compress_size, 1)) for info in infos
        )
        limits = XlsxResourceLimits(
            max_zip_members=len(infos),
            max_total_uncompressed_bytes=sum(info.file_size for info in infos),
            max_member_uncompressed_bytes=max(info.file_size for info in infos),
            max_compression_ratio=maximum_ratio,
            max_worksheet_rows=2,
            max_worksheet_cells=16,
        )

    assert len(TaxMasterXlsxAdapter(payload, amount_scale=2, limits=limits).parse()) == 1


def test_formula_and_value_workbooks_are_opened_sequentially_and_always_closed(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _xlsx()
    real_load = xlsx_module.load_workbook
    state = {"open": 0, "maximum": 0, "closed": 0}

    class TrackingWorkbook:
        def __init__(self, workbook: object) -> None:
            self._workbook = workbook
            state["open"] += 1
            state["maximum"] = max(state["maximum"], state["open"])

        def __getattr__(self, name: str):
            return getattr(self._workbook, name)

        def __getitem__(self, name: str):
            return self._workbook[name]  # type: ignore[index]

        def close(self) -> None:
            self._workbook.close()  # type: ignore[attr-defined]
            state["open"] -= 1
            state["closed"] += 1

    def tracked_load(*args: object, **kwargs: object) -> TrackingWorkbook:
        assert kwargs["read_only"] is True
        assert kwargs["keep_links"] is False
        return TrackingWorkbook(real_load(*args, **kwargs))

    monkeypatch.setattr(xlsx_module, "load_workbook", tracked_load)

    assert len(TaxMasterXlsxAdapter(payload, amount_scale=2).parse()) == 1
    assert state == {"open": 0, "maximum": 1, "closed": 2}


def test_default_limits_accept_more_than_one_hundred_companies() -> None:
    parsed = TaxMasterXlsxAdapter(_xlsx(row_count=101), amount_scale=2).parse()

    assert len(parsed) == 101
